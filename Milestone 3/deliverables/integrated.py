from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pymysql
import pandas as pd
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from datetime import datetime
import re


mydb = pymysql.connect(
  host="localhost",
  user="root",
  password="dareen123",
    database="PLDB"
)
mycursor2 = mydb.cursor()
mycursor2.execute('Use PLDB')

sql_club = "INSERT INTO Club (Name,HomeStadium,Website) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE Name=VALUES(Name),HomeStadium=VALUES(HomeStadium),Website=VALUES(Website);"
sql_players = "INSERT INTO Player (FName,LName,HomeTeam,DoB,Nationality,Position,Weight,Height) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE FName=VALUES(FName),LName=VALUES(LName),HomeTeam=VALUES(HomeTeam),DoB=VALUES(DoB),Nationality=VALUES(Nationality),Position=VALUES(Position),Weight=VALUES(Weight),Height=VALUES(Height);"
sql_playsin = "INSERT INTO PlaysIn (PlayerFName,PlayerLName,ClubName,Season) VALUES (%s, %s, %s, %s) ON DUPLICATE KEY UPDATE PlayerFName=VALUES(PlayerFName),PlayerLName=VALUES(PlayerLName),ClubName=VALUES(ClubName),Season=VALUES(Season);"
sql_stadium = "INSERT INTO Stadium (Name,ClubName,Attendance,BuildingDate,Street,District,City,Capacity,PitchSize) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE Name=VALUES(Name),ClubName=VALUES(ClubName),Attendance=VALUES(Attendance),BuildingDate=VALUES(BuildingDate),Street=VALUES(Street),District=VALUES(District),City=VALUES(City),Capacity=VALUES(Capacity),PitchSize=VALUES(PitchSize);"
sql_matches="INSERT INTO Matchs(AwayTeamScore,HomeTeamScore,Date,Season,StadiumName,HomeTeam,AwayTeam) VALUES (%s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE AwayTeamScore=VALUES(AwayTeamScore),HomeTeamScore=VALUES(HomeTeamScore),Date=VALUES(Date),Season=VALUES(Season),StadiumName=VALUES(StadiumName),HomeTeam=VALUES(HomeTeam),AwayTeam=VALUES(AwayTeam);"
sql_get= "INSERT INTO gets(ClubName,Possession,Ycards,Rcards,Goals,Shots,Fouls,MatchDate) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE ClubName=VALUES(ClubName),Possession=VALUES(Possession),Ycards=VALUES(Ycards),Rcards=VALUES(Rcards),Goals=VALUES(Goals),Shots=VALUES(Shots),Fouls=VALUES(Fouls),MatchDate=VALUES(MatchDate);"
sql_getreview= "INSERT INTO givesReviews (Username ,Rating,Text,Review_Date,Home_team,Away_Team,Match_Date) VALUES (%s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE Username=VALUES(Username),Rating=VALUES(Rating),Text=VALUES(Text),Review_Date=VALUES(Review_Date),Home_team=VALUES(Home_team),Away_Team=VALUES(Away_Team),Match_Date=VALUES(Match_Date);"


def cleanhtml(raw_html):
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext
def cleantext(raw_html):
    cleanr = re.compile('[a-zA-Z%]+')
    cleantex = re.sub(cleanr, ' ', raw_html)
    return cleantex
driver1 = webdriver.Chrome(ChromeDriverManager().install())
driver1.maximize_window()
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://www.premierleague.com/")
driver.maximize_window()

def clubs():
    driver.get("https://www.premierleague.com/")
    driver.maximize_window()

    driver.get("https://www.premierleague.com/clubs")
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(1)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass
    club_data=[]

    for i in driver.find_elements_by_xpath("//tbody[@class='allTimeDataContainer']/tr"):
        try:
            club_name=i.find_element_by_class_name('clubName').text
            club_stadium=i.find_element_by_class_name('venue').text
            driver1.get(i.find_element_by_class_name('team').find_element_by_tag_name('a').get_attribute('href'))
            driver1.implicitly_wait(30)
            time.sleep(1)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
            if 'Official Website' in driver1.find_element_by_tag_name('body').text:
                club_website=driver1.find_element_by_class_name('website').find_element_by_tag_name('a').get_attribute('href')
            else:
                club_website='NULL'
            print('Club Name: ',club_name)
            print('Club Stadium: ',club_stadium)
            print('Club Website: ',club_website)
            val_club=(club_name,club_stadium, club_website)
            val_club=tuple(None if x == 'NULL' else x for x in val_club)
            mycursor2.execute(sql_club, val_club)
            mydb.commit()
            club_data.append([club_name,club_stadium,club_website])
            print('*******************************************')
        except Exception as e:
            print('Some error might occured!')
            print(e)
            print('*******************************************')
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title='Club Data'
    sheet["A1"] = "Club Name"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Club Stadium"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Club Website"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(club_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]

    workbook.save(f"Club.csv")
    # read_file = pd.read_excel (r'./Club.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\Club.csv', index = None, header=True)
def stadiums():
    driver.get("https://www.premierleague.com/clubs")
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(1)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass
    stadium_data=[]
    stadium_links=[]
    unopened_stadiums=[]
    for i in driver.find_element_by_class_name('indexSection').find_elements_by_tag_name('li'):
        club_name=i.find_element_by_class_name('clubName').get_attribute('innerHTML')
        stadium_links.append([i.find_element_by_tag_name('a').get_attribute('href').replace('overview','stadium'),club_name])
    for i in stadium_links:
        try:
            driver1.get(i[0])
            time.sleep(2)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
            time.sleep(1)
            driver1.find_element_by_xpath("//li[contains(text(),'Stadium Information')]").click()
            stadium_name='NULL'
            club_name='NULL'
            attendance='NULL'
            built='NULL'
            street_address='NULL'
            district='NULL'
            city='NULL'
            capacity='NULL'
            pitch='NULL'
            for j in driver1.find_element_by_xpath("//div[@class='articleTab active']").text.splitlines():
                print(j)
                if ('Capacity:' in j) or ('capacity:' in j):
                    capacity=int(''.join(i for i in (j.split(':')[-1]).strip() if i.isdigit()))
                if 'attendance:' in j.lower():
                    print(j.split(' v ')[0])
                    attendance=int((''.join(i for i in j.split(' v ')[0] if i.isdigit())))
                if (('Built' in j) or ('Opened' in j)) and (len(j)<20):
                    built=int((j.split(':')[-1]).strip())
                if 'Pitch' in j:
                    pitch=j.split(':')[-1].strip()
                if 'address:' in j.lower():
                    parts=j.split(':')[-1].split(',')
                    if len(parts)==4:
                        street_address=(j.split(':')[-1].split(',')[0]+','+j.split(':')[-1].split(',')[1]).strip()
                        city=j.split(':')[-1].split(',')[2].strip()
                        district=j.split(':')[-1].split(',')[-1].strip()   
                    elif len(parts)==3:
                        street_address=(j.split(':')[-1].split(',')[0]).strip()
                        city=j.split(':')[-1].split(',')[1].strip()
                        district=j.split(':')[-1].split(',')[-1].strip()
                    else:
                        district=j.split(':')[-1].split(',')[-1].strip()
                        city=j.split(':')[-1].split(',')[-2].strip()
                        street_address=','.join(j.split(':')[-1].split(',')[:-2])
            stadium_name=driver1.find_element_by_class_name('stadiumName').text
            club_name=i[1]
            print('Stadium Name: ',stadium_name)
            print('Club Name: ',club_name)
            print('Attendance: ',attendance)    
            print('Built Date: ',built)
            print('Street Address: ',street_address)
            print('District: ',district)
            print('City: ',city)
            print('Capacity: ',capacity)
            print('Pitch: ',pitch)
            val_stadium=(stadium_name,club_name,attendance,built,street_address,district,city,capacity,pitch)
            val_stadium=tuple(None if x == 'NULL' else x for x in val_stadium)
            mycursor2.execute(sql_stadium, val_stadium)
            mydb.commit()
            stadium_data.append([stadium_name,club_name,attendance,built,street_address,district,city,capacity,pitch])
            print('**************************************')
        except Exception as e:
            print('Some error might have occured!')
            print(e)
            unopened_stadiums.append(i)
            print('**************************************')
    for i in unopened_stadiums:
        try:
            driver1.get(i[0])
            time.sleep(2)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
            time.sleep(1)
            driver1.find_element_by_xpath("//li[contains(text(),'Stadium Information')]").click()
            stadium_name='NULL'
            club_name='NULL'
            attendance='NULL'
            built='NULL'
            street_address='NULL'
            district='NULL'
            city='NULL'
            capacity='NULL'
            pitch='NULL'
            for j in driver1.find_element_by_xpath("//div[@class='articleTab active']").text.splitlines():
                print(j)
                if ('Capacity:' in j) or ('capacity:' in j):
                    capacity=int(''.join(i for i in (j.split(':')[-1]).strip() if i.isdigit()))
                if 'attendance:' in j.lower():
                    print(j.split(' v ')[0])
                    attendance=int((''.join(i for i in j.split(' v ')[0] if i.isdigit())))
                if (('Built' in j) or ('Opened' in j)) and (len(j)<20):
                    built=int((j.split(':')[-1]).strip())
                if 'Pitch' in j:
                    pitch=j.split(':')[-1].strip()
                if 'address:' in j.lower():
                    parts=j.split(':')[-1].split(',')
                    if len(parts)==4:
                        street_address=(j.split(':')[-1].split(',')[0]+','+j.split(':')[-1].split(',')[1]).strip()
                        city=j.split(':')[-1].split(',')[2].strip()
                        district=j.split(':')[-1].split(',')[-1].strip()   
                    elif len(parts)==3:
                        street_address=(j.split(':')[-1].split(',')[0]).strip()
                        city=j.split(':')[-1].split(',')[1].strip()
                        district=j.split(':')[-1].split(',')[-1].strip()
                    else:
                        district=j.split(':')[-1].split(',')[-1].strip()
                        city=j.split(':')[-1].split(',')[-2].strip()
                        street_address=','.join(j.split(':')[-1].split(',')[:-2])
            stadium_name=driver1.find_element_by_class_name('stadiumName').text
            club_name=i[1]
            print('Stadium Name: ',stadium_name)
            print('Club Name: ',club_name)
            print('Attendance: ',attendance)    
            print('Built Date: ',built)
            print('Street Address: ',street_address)
            print('District: ',district)
            print('City: ',city)
            print('Capacity: ',capacity)
            print('Pitch: ',pitch)
            val_stadium=(stadium_name,club_name,attendance,built,street_address,district,city,capacity,pitch)
            val_stadium=tuple(None if x == 'NULL' else x for x in val_stadium)
            mycursor2.execute(sql_stadium, val_stadium)
            mydb.commit()
            stadium_data.append([stadium_name,club_name,attendance,built,street_address,district,city,capacity,pitch])
            print('**************************************')
        except Exception as e:
            print(e)
            pass
            
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    #Notice to creditors
    sheet.title='Stadium Data'
    sheet["A1"] = "Stadium Name"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Club Name"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Attendance"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Built Year"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Street Address"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "District"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "City"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border
    sheet["H1"] = "Capacity"
    sheet["H1"].font = bold_font
    sheet["H1"].alignment = center_aligned_text
    sheet["H1"].border = square_border
    sheet["I1"] = "Pitch"
    sheet["I1"].font = bold_font
    sheet["I1"].alignment = center_aligned_text
    sheet["I1"].border = square_border

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(stadium_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]
        sheet[f"E{p+2}"]=q[4]
        sheet[f"F{p+2}"]=q[5]
        sheet[f"G{p+2}"]=q[6]
        sheet[f"H{p+2}"]=q[7]
        sheet[f"I{p+2}"]=q[8]

    workbook.save(f"Stadium.csv")
    # read_file = pd.read_excel (r'./Stadium.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\Stadium.csv', index = None, header=True)

    
    
def matches():
    mycursor2.execute('SET FOREIGN_KEY_CHECKS=0;')
    driver.get("https://www.premierleague.com/results")
    driver.implicitly_wait(30)
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(1)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass
    matches_data=[]
    no_of_seasons=4
    currrent_season=0
    actions = ActionChains(driver)
    for counter,i in enumerate(driver.find_elements_by_xpath("//ul[@data-dropdown-list='compSeasons']/li")):
        print('New Season =================================================')
        driver.switch_to.window(driver.current_window_handle)
        driver.execute_script("window.scrollTo(0, 0);")
        driver.find_elements_by_class_name('current')[1].click()
        driver.find_elements_by_xpath("//ul[@data-dropdown-list='compSeasons']/li")[counter].click()
        season=driver.find_element_by_xpath('//div[@data-dropdown-current="compSeasons"]').text
        while True: 
            time.sleep(2)
            print(len(driver.find_elements_by_class_name('matchFixtureContainer')))
            actions.move_to_element(driver.find_elements_by_class_name('matchFixtureContainer')[-1]).perform()
            time.sleep(2)
            driver.execute_script("window.scrollTo(0, window.scrollY+150);")
            time.sleep(5)
            if 'Loading More' not in driver.find_element_by_tag_name('body').text:
                break
        season=driver.find_element_by_xpath('//div[@data-dropdown-current="compSeasons"]').text
        for j in driver.find_elements_by_class_name('fixtures__matches-list'):
            for k in (j.find_elements_by_class_name('matchFixtureContainer')):
                try:
                    print('Home Team: ',k.get_attribute('data-home'))
                    print('Away Team: ',k.get_attribute('data-away'))
                    print('Stadium: ',k.get_attribute('data-venue').split('<')[0].replace(',',''))
                    print('Home Team Score: ',k.find_element_by_class_name('score ').text.split('-')[0])
                    print('Away Team Score: ',k.find_element_by_class_name('score ').text.split('-')[1])
                    date_parts=datetime.strptime(' '.join(j.get_attribute('data-competition-matches-list').split()[1:]),'%d %B %Y')
                    print(f'Date:{date_parts.day}-{date_parts.month}-{date_parts.year}')
                    print('Season: ',season)
                    home_team=k.get_attribute('data-home')
                    away_team=k.get_attribute('data-away')
                    stadium=k.get_attribute('data-venue').split('<')[0].replace(',','')
                    home_team_score=int(k.find_element_by_class_name('score ').text.split('-')[0])
                    away_team_score=int(k.find_element_by_class_name('score ').text.split('-')[1])
                    date=f'{date_parts.day}-{date_parts.month}-{date_parts.year}'
                    val_match=(away_team_score,home_team_score,date_parts,season,stadium,home_team,away_team)
                    val_match=tuple(None if x == 'NULL' else x for x in val_match)
                    mycursor2.execute(sql_matches, val_match)
                    mydb.commit()
                    matches_data.append([away_team_score,home_team_score,date,season,stadium,home_team,away_team])
                    print('************************')
                except Exception as e:
                    print('Some error might have occured!')
                    print(e)
                    
            print('&&&&&&&&&&&&&&&&&&&&&&&&&')
        currrent_season=currrent_season+1
        time.sleep(2)
        if currrent_season==no_of_seasons:
            break

        
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    #Notice to creditors
    sheet.title='Matches Data'
    sheet["A1"] = "Away Team Score"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Home Team Score"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Date"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Season"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Stadium"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "Home Team"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "Away Team"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border


    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(matches_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]
        sheet[f"E{p+2}"]=q[4]
        sheet[f"F{p+2}"]=q[5]
        sheet[f"G{p+2}"]=q[6]

    workbook.save(f"Matches.csv")
    # read_file = pd.read_excel (r'./Matches.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\Matches.csv', index = None, header=True)
    
def playersin():
    playersin_data=[]
    unopened_playersin=[]
    driver.get("https://www.premierleague.com/players")
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(1)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass
    time.sleep(5)
    actions = ActionChains(driver)
    current_players=[]
    previous_players=[]

    while len(driver.find_element_by_xpath('//tbody[@class="dataContainer indexSection"]').find_elements_by_tag_name('a'))<=900:
        print(len(driver.find_element_by_xpath('//tbody[@class="dataContainer indexSection"]').find_elements_by_tag_name('a'))) 
        time.sleep(2)
        actions.move_to_element(driver.find_element_by_class_name('loader')).perform()
    for k in driver.find_elements_by_xpath('//tbody[@class="dataContainer indexSection"]/tr'):
        current_players.append(k.find_element_by_tag_name('a').get_attribute('href'))
        print(k.find_element_by_tag_name('a').get_attribute('href'))
    print('current:',len(current_players))
        

    for i in (current_players):
        try:
            print(i)
            driver1.get(i)
            driver1.implicitly_wait(30)
            time.sleep(2)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                time.sleep(1)
                try:
                    driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
                except:
                    pass
            season=driver1.find_element_by_xpath("//td[@class='season']/p").text
            club_name=driver1.find_element_by_xpath("//td[@class='team']").find_element_by_tag_name('a').get_attribute('href').split('/')[-2].replace('-',' ')
            full_name= ''.join([i for i in driver1.find_element_by_xpath("//div[@class='playerDetails']").text if not i.isdigit()]).split()
            if len(full_name)==2:
                first_name= full_name[0]
                last_name= full_name[-1]
            elif len(full_name)==1:
                first_name= full_name[0]
                last_name='NULL'
            elif len(full_name)>=2:
                first_name= full_name[0]
                last_name=' '.join(full_name[1:])
            else:
                first_name='NULL'
                last_name='NULL'

            print('First Name: ',first_name)
            print('Last Name: ',last_name)
            print('Club Name: : ',club_name)
            print('Season: ',season)
            val_playsin=(first_name,last_name,club_name,season)
            val_playsin=tuple(None if x == 'NULL' else x for x in val_playsin)
            mycursor2.execute(sql_playsin, val_playsin)
            mydb.commit()
            playersin_data.append([first_name,last_name,club_name,season])
            print('***********************************************')
        except Exception as e:
            print('Some error might have occured!')
            unopened_playersin.append(i)
            print(e)
            
    for i in unopened_playersin:
        try:
            print(i)
            driver1.get(i)
            driver1.implicitly_wait(30)
            time.sleep(2)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                time.sleep(1)
                try:
                    driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
                except:
                    pass
            season=driver1.find_element_by_xpath("//td[@class='season']/p").text
            club_name=driver1.find_element_by_xpath("//td[@class='team']").find_element_by_tag_name('a').get_attribute('href').split('/')[-2].replace('-',' ')
            full_name= ''.join([i for i in driver1.find_element_by_xpath("//div[@class='playerDetails']").text if not i.isdigit()]).split()
            if len(full_name)==2:
                first_name= full_name[0]
                last_name= full_name[-1]
            elif len(full_name)==1:
                first_name= full_name[0]
                last_name='NULL'
            elif len(full_name)>=2:
                first_name= full_name[0]
                last_name=' '.join(full_name[1:])
            else:
                first_name='NULL'
                last_name='NULL'
            print('First Name: ',first_name)
            print('Last Name: ',last_name)
            print('Club Name: : ',club_name)
            print('Season: ',season)
            val_playsin=(first_name,last_name,club_name,season)
            val_playsin=tuple(None if x == 'NULL' else x for x in val_playsin)
            mycursor2.execute(sql_playsin, val_playsin)
            mydb.commit()
            playersin_data.append([first_name,last_name,club_name,season])
            print('***********************************************')
        except Exception as e:
            print(e)
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    #Notice to creditors
    sheet.title='PlayersIn Data'
    sheet["A1"] = "First Name"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Last Name"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Club Name"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Season"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(playersin_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]

    workbook.save(f"PlayersIn.csv")
    # read_file = pd.read_excel (r'./PlayersIn.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\PlayersIn.csv', index = None, header=True)

def players():
    players_data=[]
    unopened_players=[]
    driver.get("https://www.premierleague.com/players")
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(1)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass
    time.sleep(5)
    actions = ActionChains(driver)
    current_players=[]
    while True:
        print(len(driver.find_element_by_xpath('//tbody[@class="dataContainer indexSection"]').find_elements_by_tag_name('a'))) 
        time.sleep(2)
        try:
            actions.move_to_element(driver.find_element_by_class_name('loader')).perform()
        except:
            break
    for k in driver.find_elements_by_xpath('//tbody[@class="dataContainer indexSection"]/tr'):
        current_players.append([k.find_element_by_tag_name('a').get_attribute('href'),k.find_elements_by_tag_name('td')[1].text])
        print(k.find_element_by_tag_name('a').get_attribute('href'))
        print(k.find_elements_by_tag_name('td')[1].text)
    print('current:',len(current_players))
    for i in (current_players):
        try:
            print(i[0])
            driver1.get(i[0])
            driver1.implicitly_wait(30)
            time.sleep(1)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                try:
                    driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
                except:
                    pass
            team=driver1.find_element_by_xpath("//td[@class='team']").find_element_by_tag_name('a').get_attribute('href').split('/')[-2].replace('-',' ')
            full_name= ''.join([l for l in driver1.find_element_by_xpath("//div[@class='playerDetails']").text if not l.isdigit()]).split()
            if len(full_name)==2:
                first_name= full_name[0]
                last_name= full_name[-1]
            elif len(full_name)==1:
                first_name= full_name[0]
                last_name='NULL'
            elif len(full_name)>=2:
                first_name= full_name[0]
                last_name=' '.join(full_name[1:])
            else:
                first_name='NULL'
                last_name='NULL'
            if 'Date of Birth' in driver1.find_element_by_tag_name('body').text:
                dob=driver1.find_element_by_class_name('pdcol2').find_element_by_class_name('info').text
                dobe = datetime.strptime(dob.split('(')[0].strip(), '%d/%m/%Y').date()
                dob=f'{dobe.day}-{dobe.month}-{dobe.year}'
            else:
                dobe='NULL'
                dob='NULL'
            if 'Nationality' in driver1.find_element_by_tag_name('body').text:
                nationality=driver1.find_element_by_class_name('pdcol1').find_element_by_class_name('info').text
            else:
                nationality='NULL'
            if 'Height' in driver1.find_element_by_tag_name('body').text:

                height=driver1.find_element_by_class_name('pdcol3').find_element_by_class_name('info').text
                height = int(re.findall(r'\d+', height)[0])
            else:
                height='NULL'
            if 'Weight' in driver1.find_element_by_tag_name('body').get_attribute('innerHTML'):
                weight=driver1.find_element_by_class_name('u-hide').find_element_by_class_name('info').get_attribute('innerHTML')
                weight = int(re.findall(r'\d+', weight)[0]) 
            else:
                weight='NULL'
            if 'Position' in driver1.find_element_by_tag_name('body').text:
                position=i[-1]
            else:
                position='NULL'
            print('HomeTeam: ',team)
            print('Dob: ',dob)
            print('Nationality: ',nationality)
            print('Position: ', position)
            print('Height: ',height)
            print('Weight: ',weight)
            val_players=(first_name,last_name,team,dobe,nationality,position,height,weight)
            val_players=tuple(None if x == 'NULL' else x for x in val_players)
            mycursor2.execute(sql_players, val_players)
            mydb.commit()
            players_data.append([first_name,last_name,team,dob,nationality,position,height,weight])
            print('***********************************************')
        except Exception as e:
            print('Some error might have occured!')
            unopened_players.append(i)
            print(e)
    for i in unopened_players:
        try:
            print(i[0])
            driver1.get(i[0])
            driver1.implicitly_wait(30)
            time.sleep(1)
            if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                try:
                    driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
                except:
                    pass
            team=driver1.find_element_by_xpath("//td[@class='team']").find_element_by_tag_name('a').get_attribute('href').split('/')[-2].replace('-',' ')
            full_name= ''.join([l for l in driver1.find_element_by_xpath("//div[@class='playerDetails']").text if not l.isdigit()]).split()
            if len(full_name)==2:
                first_name= full_name[0]
                last_name= full_name[-1]
            elif len(full_name)==1:
                first_name= full_name[0]
                last_name='NULL'
            elif len(full_name)>=2:
                first_name= full_name[0]
                last_name=' '.join(full_name[1:])
            else:
                first_name='NULL'
                last_name='NULL'
            if 'Date of Birth' in driver1.find_element_by_tag_name('body').text:
                dob=driver1.find_element_by_class_name('pdcol2').find_element_by_class_name('info').text
                dobe = datetime.strptime(dob.split('(')[0].strip(), '%d/%m/%Y').date()
                dob=f'{dobe.day}{dobe.month}-{dobe.year}'
            else:
                dobe='NULL'
                dob='NULL'
            if 'Nationality' in driver1.find_element_by_tag_name('body').text:
                nationality=driver1.find_element_by_class_name('pdcol1').find_element_by_class_name('info').text
            else:
                nationality='NULL'
            if 'Height' in driver1.find_element_by_tag_name('body').text:

                height=driver1.find_element_by_class_name('pdcol3').find_element_by_class_name('info').text
                height = int(re.findall(r'\d+', height)[0])
            else:
                height='NULL'
            if 'Weight' in driver1.find_element_by_tag_name('body').get_attribute('innerHTML'):
                weight=driver1.find_element_by_class_name('u-hide').find_element_by_class_name('info').get_attribute('innerHTML')
                weight = int(re.findall(r'\d+', weight)[0]) 
            else:
                weight='NULL'
            if 'Position' in driver1.find_element_by_tag_name('body').text:
                position=i[-1]
            else:
                position='NULL'
            print('HomeTeam: ',team)
            print('Dob: ',dob)
            print('Nationality: ',nationality)
            print('Position: ', position)
            print('Height: ',height)
            print('Weight: ',weight)
            val_players=(first_name,last_name,team,dobe,nationality,position,height,weight)
            val_players=tuple(None if x == 'NULL' else x for x in val_players)
            mycursor2.execute(sql_players, val_players)
            mydb.commit()
            players_data.append([first_name,last_name,team,dob,nationality,position,height,weight])
            print('***********************************************')
        except Exception as e:  
            print(e)
            
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    #Notice to creditors
    sheet.title='Players Data'
    sheet["A1"] = "First Name"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Last Name"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Team"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "DoB"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Nationality"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "Position"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "Height"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border
    sheet["H1"] = "Weight"
    sheet["H1"].font = bold_font
    sheet["H1"].alignment = center_aligned_text
    sheet["H1"].border = square_border

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(players_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]
        sheet[f"E{p+2}"]=q[4]
        sheet[f"F{p+2}"]=q[5]
        sheet[f"G{p+2}"]=q[6]
        sheet[f"H{p+2}"]=q[7]

    workbook.save(f"Players.csv")

    # read_file = pd.read_excel (r'./Players.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\Players.csv', index = None, header=True)


def gets():
    mycursor2.execute('SET FOREIGN_KEY_CHECKS=0;')
    driver.get("https://www.premierleague.com/results")
    driver.implicitly_wait(30)
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(2)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass


    time.sleep(2)
    get_data=[]
    unopened_link=[]
    no_of_seasons=1
    currrent_season=0
    actions = ActionChains(driver)
    for counter,i in enumerate(driver.find_elements_by_xpath("//ul[@data-dropdown-list='compSeasons']/li")):
        print('New Season =================================================')
        driver.switch_to.window(driver.current_window_handle)
        driver.execute_script("window.scrollTo(0, 0);")
        try:
            driver.find_elements_by_class_name('current')[1].click()
        except:
            try:
                time.sleep(2)
                driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
                driver.find_elements_by_class_name('current')[1].click() 
            except:
                pass
            
        driver.find_elements_by_xpath("//ul[@data-dropdown-list='compSeasons']/li")[1].click()
        season=driver.find_element_by_xpath('//div[@data-dropdown-current="compSeasons"]').text
        try:
            while True: 
                time.sleep(2)
                driver.switch_to.window(driver.current_window_handle)
                print(len(driver.find_elements_by_class_name('matchFixtureContainer')))
                actions.move_to_element(driver.find_elements_by_class_name('matchFixtureContainer')[-1]).perform()
                driver.execute_script("window.scrollTo(0, window.scrollY+100);")
                time.sleep(5)
                if 'Loading More' not in driver.find_element_by_tag_name('body').text:
                    break
        except:
            try:
                driver.switch_to.window(driver.current_window_handle)
                while True: 
                    time.sleep(2)
                    print(len(driver.find_elements_by_class_name('matchFixtureContainer')))
                    actions.move_to_element(driver.find_elements_by_class_name('matchFixtureContainer')[-1]).perform()
                    driver.execute_script("window.scrollTo(0, window.scrollY+100);")
                    time.sleep(5)
                    if 'Loading More' not in driver.find_element_by_tag_name('body').text:
                        break
            except:
                pass
        season=driver.find_element_by_xpath('//div[@data-dropdown-current="compSeasons"]').text
        for j in driver.find_elements_by_class_name('fixtures__matches-list'):
            for k in (j.find_elements_by_class_name('matchFixtureContainer')):
                try:
                    link=k.find_element_by_class_name("postMatch").get_attribute('data-href')
                    date_parts=datetime.strptime(' '.join(j.get_attribute('data-competition-matches-list').split()[1:]),'%d %B %Y')
                    home_team=k.get_attribute('data-home')
                    away_team=k.get_attribute('data-away')
                    stadium=k.get_attribute('data-venue').split('<')[0].replace(',','')
                    home_team_score=int(k.find_element_by_class_name('score ').text.split('-')[0])
                    away_team_score=int(k.find_element_by_class_name('score ').text.split('-')[1])
                    date=f'{date_parts.day}-{date_parts.month}-{date_parts.year}'
                    print(k.find_element_by_class_name("postMatch").get_attribute('data-href'))
                    print('https:'+link)
                    driver1.get('https:'+link)
                    driver1.implicitly_wait(30)
                    driver1.switch_to.window(driver1.current_window_handle)
                    time.sleep(1)
                    if 'Cookies' in driver1.find_element_by_tag_name('body').text:
                        time.sleep(1)
                        try:
                            driver1.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
                        except:
                            pass
                    driver1.execute_script("document.getElementsByClassName('tabLinks matchNav')[0].getElementsByTagName('li')[2].click()")
                    possession=['NULL','NULL']
                    shots=['NULL','NULL']
                    ycards=['NULL','NULL']
                    rcards=['NULL','NULL']
                    foul=['NULL','NULL']
                    for l in driver1.find_elements_by_xpath('//tbody[@class="matchCentreStatsContainer"]/tr'):
                        if 'possession' in cleanhtml(l.get_attribute('innerHTML')).lower(): 
                            possession=cleantext(cleanhtml(l.get_attribute('innerHTML'))).split()
                        if ('shots' in cleanhtml(l.get_attribute('innerHTML')).lower()) and (('on target' not in cleanhtml(l.get_attribute('innerHTML')).lower())) : 
                            shots=cleantext(cleanhtml(l.get_attribute('innerHTML'))).split()
                        if 'yellow' in cleanhtml(l.get_attribute('innerHTML')).lower(): 
                            ycards=cleantext(cleanhtml(l.get_attribute('innerHTML'))).split()

                        if 'foul' in cleanhtml(l.get_attribute('innerHTML')).lower(): 
                            foul=cleantext(cleanhtml(l.get_attribute('innerHTML'))).split()

                        if 'red' in cleanhtml(l.get_attribute('innerHTML')).lower(): 
                            rcards=cleantext(cleanhtml(l.get_attribute('innerHTML'))).split()

                    try:
                        home_team_possession=float(possession[0])
                    except:
                        home_team_possession=possession[0]
                    try:
                        away_team_possession=float(possession[1])
                    except:
                        away_team_possession=possession[1]
                    try:
                        home_team_shots=int(shots[0])
                    except:
                        home_team_shots=shots[0]
                    try:
                        away_team_shots=int(shots[1]) 
                    except:
                        away_team_shots=shots[1] 
                    try:
                        home_team_ycards=int(ycards[0]) 
                    except:
                        home_team_ycards=ycards[0]
                    try:
                        away_team_ycards=int(ycards[1])
                    except:
                        away_team_ycards=ycards[1]
                    try:
                        home_team_rcards=int(rcards[0])

                    except:
                        home_team_rcards=rcards[0]
                    try:
                        away_team_rcards=int(rcards[1])
                    except:
                        away_team_rcards=rcards[1]
                    try:
                        home_team_foul=int(foul[0]) 
                    except:
                        home_team_foul=(foul[0]) 
                    try:
                        away_team_foul=int(foul[1])
                    except:
                        away_team_foul=(foul[1])


                    print('HOME TEAM')
                    print('Home Team: ',k.get_attribute('data-home'))
                    print('Stadium: ',k.get_attribute('data-venue').split('<')[0].replace(',',''))
                    print('Home Team Score: ',k.find_element_by_class_name('score ').text.split('-')[0])
                    print(f'Date:{date_parts.day}-{date_parts.month}-{date_parts.year}')
                    print('Possession: ',home_team_possession)
                    print('Shots: ',home_team_shots)
                    print('Yellow Cards: ',home_team_ycards)
                    print('Red Cards: ',home_team_rcards)
                    print('Fouls: ',home_team_foul)
                    print('Season: ',season)

                    print('AWAY TEAM')
                    print('Away Team: ',k.get_attribute('data-away'))
                    print('Stadium: ',k.get_attribute('data-venue').split('<')[0].replace(',',''))
                    print('Away Team Score: ',k.find_element_by_class_name('score ').text.split('-')[1])
                    print(f'Date:{date_parts.day}-{date_parts.month}-{date_parts.year}')
                    print('Possession: ',away_team_possession)
                    print('Shots: ',away_team_shots)
                    print('Yellow Cards: ',away_team_ycards)
                    print('Red Cards: ',away_team_rcards)
                    print('Fouls: ',away_team_foul)
                    print('Season: ',season)

                    val_get=(home_team,home_team_possession,home_team_ycards,home_team_rcards,home_team_score,home_team_shots,home_team_foul,home_team,away_team,date_parts)
                    val_get=tuple(None if x == 'NULL' else x for x in val_get)
                    print(val_get)
                    mycursor2.execute(sql_get,val_get)
                    
                    val_get=(away_team,away_team_possession,away_team_ycards,away_team_rcards,away_team_score,away_team_shots,away_team_foul,home_team,away_team,date_parts)
                    val_get=tuple(None if x == 'NULL' else x for x in val_get)
                    print(val_get)
                    mycursor2.execute(sql_get,val_get)
                    mydb.commit()
                    
                    get_data.append([home_team,home_team_possession,home_team_ycards,home_team_rcards,home_team_score,home_team_shots,home_team_foul,home_team,away_team,date])
                    get_data.append([away_team,away_team_possession,away_team_ycards,away_team_rcards,away_team_score,away_team_shots,away_team_foul,home_team,away_team,date])
                    print('************************')
                except Exception as e:
                    print('Some error occured trying again...')
                    print(e)
                    try:
                        link=k.find_element_by_class_name("postMatch").get_attribute('data-href')
                        unopened_link.append(link)
                    except Exception as e:
                        print('Some error occured!')
                        print(e)
                        pass
                    
                print('&&&&&&&&&&&&&&&&&&&&&&&&&')
            
        currrent_season=currrent_season+1
        time.sleep(2)
        if currrent_season==no_of_seasons:
            break
    print('Following matches were unable to be scraped:')
    for i in unopened_link:
        print(i)
        
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title='Get Data'
    sheet["A1"] = "Club Name"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Possession(%)"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Yellow Cards"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Red Cards"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Goals"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "Shots"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "Fouls"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border
    sheet["H1"] = "Home Team"
    sheet["H1"].font = bold_font
    sheet["H1"].alignment = center_aligned_text
    sheet["H1"].border = square_border
    sheet["I1"] = "Away Team"
    sheet["I1"].font = bold_font
    sheet["I1"].alignment = center_aligned_text
    sheet["I1"].border = square_border
    sheet["J1"] = "Match Date"
    sheet["J1"].font = bold_font
    sheet["J1"].alignment = center_aligned_text
    sheet["J1"].border = square_border


    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(get_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]
        sheet[f"E{p+2}"]=q[4]
        sheet[f"F{p+2}"]=q[5]
        sheet[f"G{p+2}"]=q[6]
        sheet[f"H{p+2}"]=q[7]
        sheet[f"I{p+2}"]=q[8]
        sheet[f"J{p+2}"]=q[9]

    workbook.save(f"gets.csv")
    # read_file = pd.read_excel (r'./gets.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\gets.csv', index = None, header=True)
def giveReviews():
    mycursor2.execute('SET FOREIGN_KEY_CHECKS=0;')
    try:
        df = pd.read_csv('givesReviews.csv')
    except:
        print('CSV file not found!')
        raise Exception('CSV file not found!')
    vals = df.values
    exceler=(vals.tolist())
    driver.get("https://www.premierleague.com/results")
    driver.implicitly_wait(30)
    time.sleep(2)
    if 'Cookies' in driver.find_element_by_tag_name('body').text:
        time.sleep(1)
        try:
            driver.find_element_by_xpath("//button[contains(text(),'Cookies')]").click()
        except:
            pass
    givesReview_data=[]
    actions = ActionChains(driver)
    no_of_records=len(df)
    exceed=False
    for counter,i in enumerate(driver.find_elements_by_xpath("//ul[@data-dropdown-list='compSeasons']/li")):
        print('New Season =================================================')
        driver.switch_to.window(driver.current_window_handle)
        driver.execute_script("window.scrollTo(0, 0);")
        driver.find_elements_by_class_name('current')[1].click()
        driver.find_elements_by_xpath("//ul[@data-dropdown-list='compSeasons']/li")[counter].click()
        
        while True: 
            time.sleep(2)
            print(len(driver.find_elements_by_class_name('matchFixtureContainer')))
            actions.move_to_element(driver.find_elements_by_class_name('matchFixtureContainer')[-1]).perform()
            time.sleep(2)
            driver.execute_script("window.scrollTo(0, window.scrollY+150);")
            time.sleep(5)
            if 'Loading More' not in driver.find_element_by_tag_name('body').text:
                break
            if len(driver.find_elements_by_class_name('matchFixtureContainer')) >=len(exceler):
                break
        for j in driver.find_elements_by_class_name('fixtures__matches-list'):
            for k in (j.find_elements_by_class_name('matchFixtureContainer')):
                try:
                    print('Home Team: ',k.get_attribute('data-home'))
                    print('Away Team: ',k.get_attribute('data-away'))
                    date_parts=datetime.strptime(' '.join(j.get_attribute('data-competition-matches-list').split()[1:]),'%d %B %Y')
                    print(f'Date:{date_parts.day}-{date_parts.month}-{date_parts.year}')
                    home_team=k.get_attribute('data-home')
                    away_team=k.get_attribute('data-away')
                    date=f'{date_parts.day}-{date_parts.month}-{date_parts.year}'
                    givesReview_data.append([home_team,away_team,date])
                    if len(givesReview_data)>=no_of_records:
                        exceed=True
                        break
                except Exception as e:
                    print('Some error might occurred trying again')
                    print(e)
            if exceed==True:
                break
        if exceed==True:
                break


    for counter,i in enumerate(vals):
        i[4]=matches_data[counter][0]
        i[5]=matches_data[counter][1]
        i[6]=matches_data[counter][2]
        
    for i in vals:
        print(i[3])
        i[3]=datetime.strptime(i[3],'%m/%d/%Y')
    for i in vals:
        print(i[6])
        i[6]=datetime.strptime(i[6],'%d-%m-%Y')
        
    for i in vals:
        val_getreview=tuple(i)
        print(val_getreview)
        val_getreview=tuple(None if x == 'NULL' else x for x in val_getreview)
        mycursor2.execute(sql_getreview, val_getreview)
        mydb.commit()
        
        

    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active

    sheet.title='getReview Data'
    sheet["A1"] = "Username"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Rating"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = 'Text Review'
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Review Date"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Home Team"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "Away Team"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "Match Date"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border


    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(exceler):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]

    for p,q in enumerate(givesReview_data):
        sheet[f"E{p+2}"]=q[0]
        sheet[f"F{p+2}"]=q[1]
        sheet[f"G{p+2}"]=q[2]

    workbook.save(f"giveReview(by_Script).csv")
    # read_file = pd.read_excel (r'./getsReview.xlsx')
    # read_file.to_csv (r'UTF-8 csvs\getsReview.csv', index = None, header=True)
def fan():
    mycursor2.execute('SET FOREIGN_KEY_CHECKS=0;')

    try:
        df = pd.read_csv('Fan.csv')
    except:
        print('CSV file not found!')
        raise Exception('CSV file not found!')
    vals = df.values
    print(vals.tolist())
    sql_fan= "INSERT INTO Fan (Username ,Email,Gender,DoB,ClubName) VALUES (%s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE Username=VALUES(Username),Email=VALUES(Email),Gender=VALUES(Gender),DoB=VALUES(DoB),ClubName=VALUES(ClubName);"
    for i in vals:
        print(' '.join(i[3].split(',')[1:]))
        i[3]=datetime.strptime(' '.join(i[3].split(',')[1:]).strip(),'%B %d  %Y')
        print(i[3])

    for i in vals:
        val_fan=tuple(i)
        print(val_fan)
        val_fan=tuple(None if x == 'NULL' else x for x in val_fan)
        mycursor2.execute(sql_fan, val_fan)
        mydb.commit()
clubs()
driver.quit()
driver1.quit()
print('Done!')
